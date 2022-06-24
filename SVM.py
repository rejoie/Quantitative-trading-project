from sklearn.svm import SVC
from sklearn.preprocessing import MinMaxScaler
from sklearn.model_selection import train_test_split
from openpyxl import load_workbook
from collections import Counter
from imblearn.over_sampling import SMOTE  # 过采样
from sklearn.metrics import confusion_matrix
import matplotlib.pyplot as plt
from sklearn.metrics import roc_auc_score, plot_roc_curve, precision_score, recall_score, f1_score
import numpy as np

pathTrain = './小组作业数据-2021.xlsx'
dataset = []
X = []
y = []
feature_index = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 16, 19, 21, 23, 24]

wb = load_workbook(pathTrain)
ws = wb['训练和测试样本集合']
for i in list(ws.rows):
    data = [j.value for j in i]
    del (data[0])
    X.append([data[1:][j] for j in feature_index])
    y.append(data[0])

labels = X[0]

X = X[1:len(X) - 2]
y = y[1:len(y) - 2]

X = MinMaxScaler().fit_transform(np.array(X))

print('数据分布：', Counter(y))

smo = SMOTE(random_state=11)
X, y = smo.fit_resample(X, y)

print('过采样数据分布：', Counter(y))

train_X, test_X, train_y, test_y = train_test_split(X, y, test_size=0.25, random_state=5)

ws = wb['预测样本']
for i in list(ws.rows):
    dataset.append([[j.value for j in i][2:][j] for j in feature_index])

del (dataset[0])

dataset = MinMaxScaler().fit_transform(np.array(dataset))

clf = SVC(gamma=0.01, C=100)
clf = clf.fit(train_X, train_y)
predict_y = clf.predict(test_X)

auc = roc_auc_score(test_y, predict_y)

print('测试集平均准确率：', clf.score(test_X, test_y))
print('AUC:', auc)
# 准确率
print('精确率: %.4f' % precision_score(y_true=test_y, y_pred=predict_y))
# 召回率
print('召回率: %.4f' % recall_score(y_true=test_y, y_pred=predict_y))
# F1
print('F1: %.4f' % f1_score(y_true=test_y, y_pred=predict_y))

# ROC曲线
plot_roc_curve(clf, test_X, test_y)
plt.title('ROC_curve' + '(AUC: ' + str(auc) + ')')
plt.ylabel('True Positive Rate')
plt.xlabel('False Positive Rate')
plt.show()

confmat = confusion_matrix(y_true=test_y, y_pred=predict_y)
print(confmat)

# 将混淆矩阵可视化
fig, ax = plt.subplots(figsize=(2.5, 2.5))
ax.matshow(confmat, cmap=plt.cm.Blues, alpha=0.3)
for i in range(confmat.shape[0]):
    for j in range(confmat.shape[1]):
        ax.text(x=j, y=i, s=confmat[i, j], va='center', ha='center')

plt.xlabel('Predict')
plt.ylabel('True')
plt.show()

row = 2
for i in clf.predict(dataset):
    ws.cell(row=row, column=2).value = i
    row = row + 1

wb.save(pathTrain)

wb = load_workbook('./预测结果.xlsx')
ws = wb['Sheet1']
row = 2
for i in clf.predict(dataset):
    ws.cell(row=row, column=3).value = i
    row = row + 1

wb.save('./预测结果.xlsx')
