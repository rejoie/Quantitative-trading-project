from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import roc_auc_score, plot_roc_curve, precision_score, recall_score, f1_score
from sklearn.tree import export_graphviz
from openpyxl import load_workbook
from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix
from imblearn.over_sampling import SMOTE  # 过采样
from collections import Counter
from matplotlib import font_manager
import matplotlib.pyplot as plt
import numpy as np
import os, re

# graphviz的安装路径
os.environ["PATH"] += os.pathsep + r'D:\Grapviz\bin'
pathTrain = './小组作业数据-2021.xlsx'
dataset = []
X = []
y = []

wb = load_workbook(pathTrain)
ws = wb['训练和测试样本集合']
for i in list(ws.rows):
    data = [j.value for j in i]
    del (data[0])
    X.append(data[1:])
    y.append(data[0])

labels = X[0]

X = X[1:len(X) - 2]
y = y[1:len(y) - 2]

X = MinMaxScaler().fit_transform(np.array(X))

print('数据分布：', Counter(y))

smo = SMOTE(random_state=11)
X, y = smo.fit_resample(X, y)

print('过采样数据分布：', Counter(y))

train_X, test_X, train_y, test_y = train_test_split(X, y, test_size=0.25, random_state=5)

ws = wb['预测样本']
for i in list(ws.rows):
    dataset.append([j.value for j in i][2:])

del (dataset[0])

dataset = MinMaxScaler().fit_transform(np.array(dataset))
## 随机森林
clf = RandomForestClassifier(n_estimators=130, max_features=None, max_depth=15, min_samples_split=20,
                             min_samples_leaf=4, random_state=0)
clf = clf.fit(train_X, train_y)
predict_y = clf.predict(test_X)

print('测试集平均准确率：', clf.score(test_X, test_y))
print('AUC:', roc_auc_score(test_y, predict_y))
# 准确率
print('精确率: %.4f' % precision_score(y_true=test_y, y_pred=predict_y))
# 召回率
print('召回率: %.4f' % recall_score(y_true=test_y, y_pred=predict_y))
# F1
print('F1: %.4f' % f1_score(y_true=test_y, y_pred=predict_y))

importances = clf.feature_importances_

indices = np.argsort(importances)[::-1]

for f in range(train_X.shape[1]):
    print("%2d) %-*s %f" % (f + 1, 30, labels[indices[f]], importances[indices[f]]))

plt.title('Feature Importance')
plt.bar(range(train_X.shape[1]), importances[indices], align='center')

plt.xticks(range(train_X.shape[1]), np.array(labels)[indices], rotation=90,
           fontproperties=font_manager.FontProperties(fname='C:\Windows\Fonts\msyh.ttc'))
plt.xlim([-1, train_X.shape[1]])
plt.tight_layout()
plt.show()

# ROC曲线
plot_roc_curve(clf, test_X, test_y)

confmat = confusion_matrix(y_true=test_y, y_pred=predict_y)
print(confmat)

# 将混淆矩阵可视化
fig, ax = plt.subplots(figsize=(2.5, 2.5))
ax.matshow(confmat, cmap=plt.cm.Blues, alpha=0.3)
for i in range(confmat.shape[0]):
    for j in range(confmat.shape[1]):
        ax.text(x=j, y=i, s=confmat[i, j], va='center', ha='center')

plt.xlabel('Predict')
plt.ylabel('True')
plt.show()

row = 2
for i in clf.predict(dataset):
    ws.cell(row=row, column=2).value = i
    row = row + 1

wb.save(pathTrain)

dot_data = export_graphviz(clf.estimators_[1], feature_names=labels, class_names=['不违约', '违约'], filled=True,
                           out_file='RF_old.dot', rounded=True)

f_old = open('RF_old.dot', 'r', encoding='utf-8')
f_new = open('RF.dot', 'w', encoding='utf-8')
for line in f_old:
    if 'fontname' in line:
        font_re = 'fontname=(.*?)]'
        old_font = re.findall(font_re, line)[0]
        line = line.replace(old_font, 'SimHei')
    f_new.write(line)
f_old.close()
f_new.close()
os.system('dot -Tpng RF.dot -o RF.png')  # dot -Tpng RF.dot -o RF.png

wb = load_workbook('./预测结果.xlsx')
ws = wb['Sheet1']
row = 2
for i in clf.predict(dataset):
    ws.cell(row=row, column=1).value = i
    row = row + 1

wb.save('./预测结果.xlsx')
